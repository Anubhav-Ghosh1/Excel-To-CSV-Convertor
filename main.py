import pandas as pd
import openpyxl
import os

def convert_excel_to_csv_with_warnings(excel_file: str, sheet_name: str, output_csv: str):
    """
        What this does?
        It takes an Excel file as input, reads the specified sheet, and converts it to a CSV file.

        How?
        - It uses the openpyxl library to read the Excel file.
        - It reads the specified sheet using the sheet_name parameter.
        - It converts the sheet data to a DataFrame using pandas.
        - It saves the DataFrame to a CSV file using the to_csv method.
    """
    try:
        # Read the Excel file
        workbook = openpyxl.load_workbook(excel_file)
        print("Workbook", workbook)

        # Check for multiple sheets
        sheet_names = workbook.sheetnames # This will return the list of sheets that are present in the excel file
        if len(sheet_names) > 1:
            print(f"Warning: The Excel file has multiple sheets: {sheet_names}. Only '{sheet_name}' will be converted.")

        # If we want to get data from specific sheet then we can provide its name
        # if sheet_name not in sheet_names:
        #     print(f"Error: Sheet '{sheet_name}' does not exist in the Excel file.")
        #     return

        sheet = workbook[sheet_name]
        print("Sheet", sheet)
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
        print("Data Frame", df)
        df.to_csv(output_csv, index=False) # If the index is true then the fill will have index values
        print(f"Excel sheet '{sheet_name}' has been successfully converted to '{output_csv}'.")

    except Exception as e:
        print(f"An error occurred: {e}")

# Example Usage
excel_file = "random_transactions.xlsx"
sheet_name = "Sheet1"
output_csv = "output2.csv"

convert_excel_to_csv_with_warnings(excel_file, sheet_name, output_csv)


def convert_excel_with_multiple_sheet_to_csv_with_warnings(excel_file: str, output_dir: str):
    """
        What this does?
        It takes an Excel file as input, reads the specified sheet, and converts it to a CSV file.

        How?
        - It uses the openpyxl library to read the Excel file.
        - It reads the specified sheet using the sheet_name parameter.
        - It converts the sheet data to a DataFrame using pandas.
        - It saves the DataFrame to a CSV file using the to_csv method.
    """
    try:
        # Read the Excel file
        workbook = openpyxl.load_workbook(excel_file)
        print("Workbook", workbook)

        # Check for multiple sheets
        sheet_names = workbook.sheetnames
        for sheet in sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet)
            output_csv = os.path.join(output_dir, f"{sheet}.csv")
            df.to_csv(output_csv, index=False)

    except Exception as e:
        print(f"An error occurred: {e}")

# Example Usage
excel_file = "random_transactions.xlsx"
output_dir = "output_csvs"  # Directory where the CSV files will be saved

# Ensure the output directory exists
os.makedirs(output_dir, exist_ok=True)
convert_excel_with_multiple_sheet_to_csv_with_warnings(excel_file, output_dir)